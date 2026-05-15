from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.simulate_registration import (
    REQUIRED_FIELDS,
    SimulatedRecord,
    _fill_missing_with_stubs,
    _missing_required_fields,
    _row_to_form,
    _write_xlsx,
)


class StubFillTests(unittest.TestCase):
    def test_fill_missing_with_stubs_keeps_existing_values(self) -> None:
        row = {
            "fio": "Иванов И.И.",
            "group_name": "",
            "workplace": "",
            "position": "методист",
            "phone": "",
            "supervisor": "",
            "report_url": "",
        }

        enriched = _fill_missing_with_stubs(row, 0)

        self.assertEqual(enriched["fio"], "Иванов И.И.")
        self.assertEqual(enriched["position"], "методист")
        self.assertTrue(enriched["group_name"])
        self.assertTrue(enriched["phone"])
        self.assertTrue(enriched["report_url"].startswith("https://docs.google.com/"))

    def test_missing_required_fields_detects_only_empty_values(self) -> None:
        row = {
            "fio": " ",
            "group_name": "МТ-1",
            "workplace": "",
            "position": "аналитик",
            "phone": "",
            "supervisor": "Петров П.П.",
            "report_url": "",
        }

        missing = _missing_required_fields(row)

        self.assertEqual(missing, ["fio", "workplace", "phone", "report_url"])


class XlsxPreviewTests(unittest.TestCase):
    def test_write_xlsx_creates_preview_sheet(self) -> None:
        real_form = _row_to_form(
            {
                "fio": "Иванов И.И.",
                "group_name": "МТ-1",
                "workplace": "Школа",
                "position": "учитель",
                "phone": "+7 700 000 00 00",
                "supervisor": "Петров П.П.",
                "report_url": "https://docs.google.com/document/d/test/edit",
            },
            telegram_id="111",
            check_links=False,
        )
        stub_form = _row_to_form(
            _fill_missing_with_stubs({"fio": "", "group_name": ""}, 1),
            telegram_id="222",
            check_links=False,
        )
        records = [
            SimulatedRecord(real_form, "real", []),
            SimulatedRecord(stub_form, "real+filled", list(REQUIRED_FIELDS)),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "preview.xlsx"
            _write_xlsx(records, path)

            wb = load_workbook(path)
            self.assertEqual(wb.sheetnames, ["Регистрация", "Предпросмотр"])

            preview = wb["Предпросмотр"]
            self.assertEqual(preview["A2"].value, "real")
            self.assertEqual(preview["A3"].value, "real+filled")
            self.assertEqual(preview["B3"].value, ", ".join(REQUIRED_FIELDS))


if __name__ == "__main__":
    unittest.main()
