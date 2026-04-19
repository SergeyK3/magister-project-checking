"""Симулятор регистрации магистрантов через бота — без реального Telegram.

Прогоняет ту же бизнес-логику (`magister_checking.bot.handlers`) поверх
in-memory FakeWorksheet и сохраняет результат как XLSX (или печатает TSV).

Источник данных:
- ``--input-csv PATH`` — CSV с колонками
  ``fio,group_name,workplace,position,phone,supervisor,report_url``
  (любая колонка может быть пустой — это эмулирует «дозаполнение позднее»).
- ``--from-sheet`` — прочитать первые N строк уже существующей Google Sheets,
  используя `.env`/Service Account (см. magister_checking.bot.config). В этом
  режиме telegram_id и т.п. берутся из ваших же данных в таблице.

Выход:
- ``--output-xlsx PATH`` — XLSX-файл с шапкой п.8.1 ТЗ и заполненными строками.
- Если не указан — печатает TSV в stdout.

Telegram ID:
- Для CSV-источника по умолчанию остаётся пустым (см. п.6.2 ТЗ — реальный Tg ID
  фиксируется только при живом /start). Можно прокинуть фейковые ID опцией
  ``--fake-telegram-id-base 1000`` — тогда строкам присвоятся 1000, 1001, …
- Для ``--from-sheet`` Tg ID берётся из соответствующей колонки исходной таблицы.

Сетевая проверка ссылок:
- По умолчанию первичная проверка URL отключена (``--no-link-check``), чтобы
  симуляция работала оффлайн и детерминированно. Для каждой непустой ссылки
  устанавливается ``report_url_valid = "yes"|"no"`` по формату, без HTTP.
- ``--check-links`` включает реальный HTTP-запрос (как в боевом боте).
"""

from __future__ import annotations

import argparse
import csv
import io
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional

from magister_checking.bot.models import (
    FIELD_LABELS,
    REQUIRED_FIELDS,
    SHEET_HEADER,
    UserForm,
    compute_fill_status,
    get_missing_field_keys,
)
from magister_checking.bot.validation import (
    check_report_url,
    is_valid_url,
    normalize_text,
)


CSV_FIELDS = (
    "fio",
    "group_name",
    "workplace",
    "position",
    "phone",
    "supervisor",
    "report_url",
)


@dataclass
class SimulatedRecord:
    form: UserForm
    source_kind: str
    missing_before_fill: List[str]


def _load_csv(path: Path) -> List[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        missing = [name for name in CSV_FIELDS if name not in (reader.fieldnames or [])]
        if missing:
            raise SystemExit(
                "В CSV отсутствуют колонки: " + ", ".join(missing)
                + "\nОжидаются: " + ", ".join(CSV_FIELDS)
            )
        return [dict(row) for row in reader]


def _load_from_sheet(limit: int) -> List[dict]:
    from magister_checking.bot.config import load_config
    from magister_checking.bot.sheets_repo import get_worksheet

    cfg = load_config()
    ws = get_worksheet(cfg)
    raw_rows = ws.get_all_values()
    if not raw_rows:
        return []
    header = raw_rows[0]
    name_to_idx = {name: idx for idx, name in enumerate(header)}
    out: List[dict] = []
    for row in raw_rows[1 : 1 + limit]:
        record = {name: (row[idx] if idx < len(row) else "") for name, idx in name_to_idx.items()}
        out.append(record)
    return out


def _stub_value(field_key: str, idx: int) -> str:
    number = idx + 1
    if field_key == "fio":
        return f"Тестовый Магистрант {number:02d}"
    if field_key == "group_name":
        return f"МТ-{26 + (idx % 3)}-{number:02d}"
    if field_key == "workplace":
        return f"Организация {number:02d}"
    if field_key == "position":
        return "специалист"
    if field_key == "phone":
        return f"+7 700 000 {number:02d} {number:02d}"
    if field_key == "supervisor":
        return f"Руководитель {number:02d}"
    if field_key == "report_url":
        return f"https://docs.google.com/document/d/stub-report-{number:02d}/edit"
    if field_key == "telegram_username":
        return f"sim_user_{number:02d}"
    if field_key == "telegram_first_name":
        return "Тестовый"
    if field_key == "telegram_last_name":
        return f"Пользователь {number:02d}"
    return ""


def _build_stub_row(idx: int) -> dict:
    row = {field: _stub_value(field, idx) for field in CSV_FIELDS}
    row["telegram_username"] = _stub_value("telegram_username", idx)
    row["telegram_first_name"] = _stub_value("telegram_first_name", idx)
    row["telegram_last_name"] = _stub_value("telegram_last_name", idx)
    return row


def _missing_required_fields(row: dict) -> List[str]:
    missing: List[str] = []
    for field_key in REQUIRED_FIELDS:
        value = normalize_text(row.get(field_key, "") or "")
        if not value:
            missing.append(field_key)
    return missing


def _fill_missing_with_stubs(row: dict, idx: int) -> dict:
    enriched = dict(row)
    for field_key in REQUIRED_FIELDS:
        if not normalize_text(enriched.get(field_key, "") or ""):
            enriched[field_key] = _stub_value(field_key, idx)
    for field_key in ("telegram_username", "telegram_first_name", "telegram_last_name"):
        if not normalize_text(enriched.get(field_key, "") or ""):
            enriched[field_key] = _stub_value(field_key, idx)
    return enriched


def _row_to_form(
    row: dict,
    *,
    telegram_id: str,
    check_links: bool,
) -> UserForm:
    """Прогоняет одну запись через те же шаги, что и handlers.receive_field."""

    form = UserForm()
    form.telegram_id = telegram_id
    form.telegram_username = row.get("telegram_username", "") or ""
    form.telegram_first_name = row.get("telegram_first_name", "") or ""
    form.telegram_last_name = row.get("telegram_last_name", "") or ""

    last_action = "start_new"
    for field_key in REQUIRED_FIELDS:
        raw_value = row.get(field_key, "") or ""
        value = normalize_text(raw_value)
        last_action = f"ask_{field_key}"
        if not value:
            last_action = f"skipped_{field_key}"
            setattr(form, field_key, "")
            if field_key == "report_url":
                form.report_url_valid = ""
                form.report_url_accessible = ""
                form.report_url_public_guess = ""
            continue
        setattr(form, field_key, value)
        last_action = f"answered_{field_key}"
        if field_key == "report_url":
            if check_links:
                valid, accessible, public = check_report_url(value)
            else:
                valid = "yes" if is_valid_url(value) else "no"
                accessible = ""
                public = ""
            form.report_url_valid = valid
            form.report_url_accessible = accessible
            form.report_url_public_guess = public

    form.fill_status = compute_fill_status(form).value
    if get_missing_field_keys(form):
        form.last_action = last_action or "show_summary"
    else:
        form.last_action = "confirmed_save"
    return form


def _print_tsv(forms: Iterable[UserForm], stream) -> None:
    writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
    writer.writerow(SHEET_HEADER)
    for form in forms:
        writer.writerow([str(getattr(form, name) or "") for name in SHEET_HEADER])


def _write_xlsx(records: List[SimulatedRecord], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise SystemExit(
            "Для XLSX-вывода нужна библиотека openpyxl: pip install openpyxl"
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрация"
    ws.append(SHEET_HEADER)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    fills = {
        "real": PatternFill("solid", fgColor="E2F0D9"),
        "real+filled": PatternFill("solid", fgColor="FFF2CC"),
        "stub": PatternFill("solid", fgColor="D9EAF7"),
    }

    for record in records:
        form = record.form
        ws.append([str(getattr(form, name) or "") for name in SHEET_HEADER])
        fill = fills.get(record.source_kind)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=len(SHEET_HEADER)).coordinate}"

    for column_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in column_cells), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 60)

    preview = wb.create_sheet("Предпросмотр")
    preview_header = [
        "source_kind",
        "missing_before_fill",
        *SHEET_HEADER,
    ]
    preview.append(preview_header)
    for cell in preview[1]:
        cell.font = header_font
        cell.fill = header_fill
    for record in records:
        form = record.form
        preview.append(
            [
                record.source_kind,
                ", ".join(record.missing_before_fill),
                *[str(getattr(form, name) or "") for name in SHEET_HEADER],
            ]
        )
        fill = fills.get(record.source_kind)
        if fill:
            for cell in preview[preview.max_row]:
                cell.fill = fill
    preview.freeze_panes = "A2"
    preview.auto_filter.ref = (
        f"A1:{preview.cell(row=preview.max_row, column=len(preview_header)).coordinate}"
    )
    for column_cells in preview.columns:
        max_len = max((len(str(c.value or "")) for c in column_cells), default=10)
        preview.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _summary(records: List[SimulatedRecord]) -> str:
    forms = [record.form for record in records]
    counts: dict[str, int] = {}
    for form in forms:
        counts[form.fill_status] = counts.get(form.fill_status, 0) + 1
    by_status = ", ".join(f"{k}={v}" for k, v in sorted(counts.items()))
    source_counts: dict[str, int] = {}
    for record in records:
        source_counts[record.source_kind] = source_counts.get(record.source_kind, 0) + 1
    by_source = ", ".join(f"{k}={v}" for k, v in sorted(source_counts.items()))
    lines = [f"Всего записей: {len(forms)}", f"Статусы: {by_status}", f"Источники: {by_source}"]
    for idx, record in enumerate(records, start=1):
        form = record.form
        missing = ", ".join(FIELD_LABELS[k] for k in get_missing_field_keys(form))
        info = f"  {idx}. {form.fio or '(без ФИО)'} — {form.fill_status} [{record.source_kind}]"
        if missing:
            info += f" [не заполнено: {missing}]"
        lines.append(info)
    return "\n".join(lines)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Симуляция регистрации магистрантов без реального Telegram."
    )
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument(
        "--input-csv",
        type=Path,
        help="CSV с колонками fio,group_name,workplace,position,phone,supervisor,report_url",
    )
    src.add_argument(
        "--from-sheet",
        action="store_true",
        help="Прочитать первые N строк уже существующей Google Sheets (использует .env / SA)",
    )
    parser.add_argument(
        "--limit", type=int, default=5, help="сколько строк взять из источника (по умолчанию 5)"
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help="путь к XLSX-результату; если не указан — TSV в stdout",
    )
    parser.add_argument(
        "--check-links",
        action="store_true",
        help="реально дернуть report_url через requests (по умолчанию выключено)",
    )
    parser.add_argument(
        "--fake-telegram-id-base",
        type=int,
        default=0,
        help="если > 0 — присвоить синтетические telegram_id начиная с этого числа",
    )
    parser.add_argument(
        "--fill-missing-with-stubs",
        action="store_true",
        help="дозаполнить пустые обязательные поля правдоподобными заглушками",
    )
    parser.add_argument(
        "--append-stub-rows",
        type=int,
        default=0,
        help="добавить N полностью синтетических строк после основного источника",
    )

    args = parser.parse_args(argv)

    if args.from_sheet:
        rows = _load_from_sheet(args.limit)
    else:
        rows = _load_csv(args.input_csv)[: args.limit]

    if not rows:
        print("Источник не содержит данных.", file=sys.stderr)
        return 1

    records: List[SimulatedRecord] = []
    for idx, row in enumerate(rows):
        missing_before_fill = _missing_required_fields(row)
        enriched_row = (
            _fill_missing_with_stubs(row, idx)
            if args.fill_missing_with_stubs
            else dict(row)
        )
        if args.from_sheet:
            telegram_id = (row.get("telegram_id") or "").strip()
        elif args.fake_telegram_id_base > 0:
            telegram_id = str(args.fake_telegram_id_base + idx)
        else:
            telegram_id = ""
        source_kind = "real+filled" if args.fill_missing_with_stubs and missing_before_fill else "real"
        records.append(
            SimulatedRecord(
                form=_row_to_form(
                    enriched_row,
                    telegram_id=telegram_id,
                    check_links=args.check_links,
                ),
                source_kind=source_kind,
                missing_before_fill=missing_before_fill,
            )
        )

    base_count = len(rows)
    for offset in range(args.append_stub_rows):
        row_idx = base_count + offset
        stub_row = _build_stub_row(row_idx)
        if args.fake_telegram_id_base > 0:
            telegram_id = str(args.fake_telegram_id_base + row_idx)
        else:
            telegram_id = ""
        records.append(
            SimulatedRecord(
                form=_row_to_form(stub_row, telegram_id=telegram_id, check_links=args.check_links),
                source_kind="stub",
                missing_before_fill=list(REQUIRED_FIELDS),
            )
        )

    if args.output_xlsx:
        _write_xlsx(records, args.output_xlsx)
        print(f"Сохранено: {args.output_xlsx}")
    else:
        _print_tsv((record.form for record in records), sys.stdout)

    print("\n" + _summary(records), file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
